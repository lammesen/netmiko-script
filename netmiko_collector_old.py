#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Netmiko Device Command Collector

This script connects to multiple Cisco IOS devices via SSH using Netmiko,
executes a list of commands on each device, and saves the output to a CSV file.

Author: Network Automation
Date: 2025-10-24
Version: 2.0.0
"""

import csv
import io
import json
import logging
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from getpass import getpass
from pathlib import Path
from typing import Dict, List, Optional

import typer
from typing_extensions import Annotated

# Set UTF-8 encoding for Windows console to support emojis
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")
    except (AttributeError, io.UnsupportedOperation):
        # If already wrapped or not supported, continue
        pass

# Constants for boolean value parsing
TRUTHY_VALUES = ("true", "yes", "1")

# Worker limits
MIN_WORKERS = 1
MAX_WORKERS = 20

# Retry configuration
MAX_RETRY_ATTEMPTS = 3
RETRY_MIN_WAIT = 1
RETRY_MAX_WAIT = 10

# Default configuration
DEFAULT_CONFIG = {
    "default_device_type": "cisco_ios",
    "strip_whitespace": True,
    "max_workers": 5,
    "connection_timeout": 30,
    "command_timeout": 60,
    "enable_session_logging": False,
    "enable_mode": False,
    "retry_on_failure": True,
}

# Configuration file location
CONFIG_FILE = Path.home() / ".netmiko_collector_config.json"

# Command database for different device types
DEVICE_COMMANDS = {
    "cisco_ios": {
        "show": [
            "show version",
            "show running-config",
            "show startup-config",
            "show ip interface brief",
            "show ip route",
            "show interfaces",
            "show interfaces status",
            "show interfaces description",
            "show vlan brief",
            "show vtp status",
            "show spanning-tree",
            "show cdp neighbors",
            "show cdp neighbors detail",
            "show lldp neighbors",
            "show mac address-table",
            "show arp",
            "show inventory",
            "show environment",
            "show logging",
            "show processes cpu",
            "show memory",
            "show users",
            "show privilege",
            "show clock",
            "show ntp associations",
            "show ip protocols",
            "show ip ospf neighbor",
            "show ip eigrp neighbors",
            "show ip bgp summary",
            "show access-lists",
            "show ip nat translations",
            "show standby",
            "show crypto isakmp sa",
            "show crypto ipsec sa",
        ],
        "filters": [
            " | include ",
            " | exclude ",
            " | begin ",
            " | section ",
        ],
    },
    "cisco_xe": {
        "show": [
            "show version",
            "show running-config",
            "show startup-config",
            "show ip interface brief",
            "show ip route",
            "show interfaces",
            "show interfaces status",
            "show interfaces description",
            "show vlan brief",
            "show vtp status",
            "show spanning-tree",
            "show cdp neighbors",
            "show cdp neighbors detail",
            "show lldp neighbors",
            "show mac address-table",
            "show arp",
            "show inventory",
            "show environment all",
            "show logging",
            "show processes cpu",
            "show memory statistics",
            "show users",
            "show platform",
            "show platform software status control-processor brief",
            "show redundancy",
            "show issu state",
            "show license summary",
            "show ip protocols",
            "show ip ospf neighbor",
            "show ip eigrp neighbors",
            "show ip bgp summary",
            "show ip bgp neighbors",
            "show access-lists",
            "show crypto isakmp sa",
            "show crypto ipsec sa",
            "show sdwan system status",
        ],
        "filters": [
            " | include ",
            " | exclude ",
            " | begin ",
            " | section ",
        ],
    },
    "cisco_nxos": {
        "show": [
            "show version",
            "show running-config",
            "show startup-config",
            "show interface brief",
            "show ip route",
            "show interface",
            "show interface status",
            "show interface description",
            "show vlan",
            "show vpc",
            "show vpc brief",
            "show spanning-tree",
            "show cdp neighbors",
            "show cdp neighbors detail",
            "show lldp neighbors",
            "show mac address-table",
            "show ip arp",
            "show inventory",
            "show environment",
            "show logging last 100",
            "show processes cpu",
            "show system resources",
            "show users",
            "show module",
            "show hardware",
            "show feature",
            "show ip ospf neighbors",
            "show ip eigrp neighbors",
            "show ip bgp summary",
            "show ip bgp neighbors",
            "show ip access-lists",
            "show hsrp brief",
            "show vrrp",
            "show port-channel summary",
        ],
        "filters": [
            " | include ",
            " | exclude ",
            " | begin ",
            " | grep ",
        ],
    },
    "cisco_asa": {
        "show": [
            "show version",
            "show running-config",
            "show startup-config",
            "show interface ip brief",
            "show route",
            "show interface",
            "show nameif",
            "show xlate",
            "show conn",
            "show nat",
            "show access-list",
            "show crypto isakmp sa",
            "show crypto ipsec sa",
            "show vpn-sessiondb",
            "show failover",
            "show logging",
            "show processes cpu-usage",
            "show memory",
            "show blocks",
            "show inventory",
            "show environment",
            "show clock",
            "show ntp associations",
        ],
        "filters": [
            " | include ",
            " | exclude ",
            " | begin ",
            " | grep ",
        ],
    },
    "arista_eos": {
        "show": [
            "show version",
            "show running-config",
            "show startup-config",
            "show ip interface brief",
            "show ip route",
            "show interfaces",
            "show interfaces status",
            "show interfaces description",
            "show vlan",
            "show spanning-tree",
            "show lldp neighbors",
            "show mac address-table",
            "show arp",
            "show inventory",
            "show environment all",
            "show logging last 100",
            "show processes top once",
            "show users",
            "show module",
            "show ip bgp summary",
            "show ip ospf neighbor",
            "show mlag",
            "show vxlan vtep",
            "show vxlan vni",
        ],
        "filters": [
            " | include ",
            " | exclude ",
            " | begin ",
            " | grep ",
        ],
    },
    "juniper_junos": {
        "show": [
            "show version",
            "show configuration",
            "show interfaces terse",
            "show interfaces descriptions",
            "show route",
            "show chassis hardware",
            "show chassis environment",
            "show system alarms",
            "show system uptime",
            "show system users",
            "show log messages",
            "show arp",
            "show lldp neighbors",
            "show ethernet-switching table",
            "show ospf neighbor",
            "show bgp summary",
            "show security policies",
            "show security ipsec security-associations",
        ],
        "filters": [
            " | match ",
            " | except ",
            " | find ",
            " | display set",
            " | display xml",
        ],
    },
}

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import (
        NetmikoAuthenticationException,
        NetmikoTimeoutException,
    )
except ImportError:
    print("Error: Netmiko is not installed. Please run: pip install -r requirements.txt")
    sys.exit(1)

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
    from rich.table import Table

    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False
    Console = None

try:
    from tqdm import tqdm

    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

    TENACITY_AVAILABLE = True
except ImportError:
    TENACITY_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from prompt_toolkit import prompt
    from prompt_toolkit.application import Application
    from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
    from prompt_toolkit.completion import FuzzyCompleter, WordCompleter
    from prompt_toolkit.formatted_text import HTML
    from prompt_toolkit.history import InMemoryHistory
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout.containers import HSplit, Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.layout.layout import Layout
    from prompt_toolkit.styles import Style

    PROMPT_TOOLKIT_AVAILABLE = True
except ImportError:
    PROMPT_TOOLKIT_AVAILABLE = False


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("netmiko_collector.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

# Initialize rich console if available
console = Console() if RICH_AVAILABLE else None


def print_banner(text: str, style: str = "bold blue") -> None:
    """Print a banner with rich formatting if available."""
    if RICH_AVAILABLE and console:
        console.print(Panel(text, style=style, expand=False))
    else:
        print("\n" + "=" * 60)
        print(text)
        print("=" * 60)


def print_success(text: str) -> None:
    """Print success message."""
    if RICH_AVAILABLE and console:
        console.print(f"✅ {text}", style="bold green")
    else:
        print(f"✅ {text}")


def print_error(text: str) -> None:
    """Print error message."""
    if RICH_AVAILABLE and console:
        console.print(f"❌ {text}", style="bold red")
    else:
        print(f"❌ {text}")


def print_warning(text: str) -> None:
    """Print warning message."""
    if RICH_AVAILABLE and console:
        console.print(f"⚠️  {text}", style="bold yellow")
    else:
        print(f"⚠️  {text}")


def print_info(text: str) -> None:
    """Print info message."""
    if RICH_AVAILABLE and console:
        console.print(f"ℹ️  {text}", style="bold cyan")
    else:
        print(f"ℹ️  {text}")


def get_positive_int_input(
    prompt: str, current_value: int, min_val: int = 1, max_val: Optional[int] = None
) -> Optional[int]:
    """
    Get and validate positive integer input from user.

    Args:
        prompt: Prompt text to display
        current_value: Current value to show as default
        min_val: Minimum allowed value
        max_val: Maximum allowed value (optional)

    Returns:
        Validated integer or None if invalid/empty
    """
    try:
        full_prompt = f"{prompt} [{current_value}]: "
        new_value = input(full_prompt).strip()
        if new_value:
            value = int(new_value)
            if value < min_val or (max_val and value > max_val):
                range_text = f"{min_val} and {max_val}" if max_val else f"at least {min_val}"
                print_error(f"Value must be between {range_text}")
                return None
            return value
    except ValueError:
        print_error("Invalid number")
    return None


def expand_path(path_str: str) -> Path:
    """
    Expand user home directory and resolve path.

    Args:
        path_str: Path string potentially containing ~

    Returns:
        Expanded Path object
    """
    return Path(path_str).expanduser().resolve()


# Custom styling for prompt_toolkit - clean and simple
CUSTOM_STYLE = (
    Style.from_dict(
        {
            # Autocomplete dropdown - minimal and clean
            "completion-menu.completion": "#888888",  # Gray text for unselected
            # Bright cyan for selected (no background)
            "completion-menu.completion.current": "#00aaff bold",
            # Scrollbar - subtle appearance
            "scrollbar.background": "bg:#2a2a2a",
            "scrollbar.button": "bg:#00aaff",
            # Prompt styling
            "prompt": "cyan bold",
            "input": "#ffffff",  # White input text
        }
    )
    if PROMPT_TOOLKIT_AVAILABLE
    else None
)


def select_from_menu(title: str, options: List[tuple], style: str = "bold cyan") -> Optional[str]:
    """
    Display an arrow-key navigable menu and return selected option.

    Args:
        title: Menu title to display
        options: List of tuples (value, description) for each menu item
        style: Rich style for the title banner

    Returns:
        Selected option value or None if cancelled (Escape pressed)
    """
    if not PROMPT_TOOLKIT_AVAILABLE:
        # Fallback to basic input
        print_banner(title, style)
        for idx, (value, desc) in enumerate(options, 1):
            print(f"{idx}. {desc}")

        while True:
            try:
                choice = input("\nEnter your choice: ").strip()
                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(options):
                        return options[idx][0]
                print_error("Invalid choice. Please try again.")
            except KeyboardInterrupt:
                return None

    # Use prompt_toolkit for beautiful arrow-key navigation
    print_banner(title, style)

    # Track current selection
    current_index = [0]  # Use list to allow modification in nested function

    def get_formatted_text():
        """Generate formatted menu text with highlighted selection."""
        result = []
        for idx, (value, desc) in enumerate(options):
            if idx == current_index[0]:
                # Highlighted item with arrow emoji
                result.append(("class:selected", f"  ▶ {desc}\n"))
            else:
                # Normal item with circle
                result.append(("", f"  ○ {desc}\n"))
        result.append(("class:help", "\n↑/↓ Navigate • Enter Select • Esc Cancel"))
        return result

    # Create key bindings
    kb = KeyBindings()
    exit_value = {"value": None}

    @kb.add("up")
    def _(event):
        current_index[0] = (current_index[0] - 1) % len(options)
        event.app.invalidate()

    @kb.add("down")
    def _(event):
        current_index[0] = (current_index[0] + 1) % len(options)
        event.app.invalidate()

    @kb.add("enter")
    def _(event):
        exit_value["value"] = options[current_index[0]][0]
        exit_value["description"] = options[current_index[0]][1]
        event.app.exit()

    @kb.add("escape")
    def _(event):
        exit_value["value"] = None
        event.app.exit()

    @kb.add("q")
    def _(event):
        exit_value["value"] = None
        event.app.exit()

    @kb.add("c-c")
    def _(event):
        exit_value["value"] = None
        event.app.exit()

    # Create layout with dynamic content
    menu_control = FormattedTextControl(text=get_formatted_text, focusable=True)

    container = HSplit(
        [
            Window(height=1),
            Window(content=menu_control, height=len(options) + 2),
        ]
    )

    # Create application
    app = Application(
        layout=Layout(container),
        key_bindings=kb,
        mouse_support=True,
        full_screen=False,
        style=Style.from_dict(
            {
                "selected": "bg:#00aaff #000000 bold",
                "help": "cyan",
            }
        ),
    )

    # Run application
    try:
        app.run()

        # Show selection confirmation if something was selected
        if exit_value["value"] is not None and "description" in exit_value:
            if RICH_AVAILABLE and console:
                console.print(f"✅ Selected: [cyan]{exit_value['description']}[/cyan]\n")
            else:
                print(f"✅ Selected: {exit_value['description']}\n")

        return exit_value["value"]
    except KeyboardInterrupt:
        return None


def open_file_in_editor(file_path: str) -> bool:
    """
    Open a file in the default system editor.

    Args:
        file_path: Path to the file to edit

    Returns:
        True if successful, False otherwise
    """
    import platform
    import subprocess  # nosec B404 - subprocess needed for cross-platform editor support

    file_path_obj = Path(file_path)

    # Create file if it doesn't exist
    if not file_path_obj.exists():
        try:
            file_path_obj.touch()
            print_info(f"Created new file: {file_path}")
        except Exception as e:
            print_error(f"Failed to create file: {e}")
            return False

    try:
        system = platform.system()

        if system == "Windows":
            # Try notepad first, then default editor
            subprocess.run(["notepad", str(file_path_obj)], check=False)  # nosec B603 B607
        elif system == "Darwin":  # macOS
            subprocess.run(["open", "-e", str(file_path_obj)], check=False)  # nosec B603 B607
        else:  # Linux
            # Try common editors
            editors = ["nano", "vim", "vi", "gedit", "kate"]
            for editor in editors:
                try:
                    subprocess.run([editor, str(file_path_obj)], check=False)  # nosec B603 B607
                    break
                except FileNotFoundError:
                    continue

        return True

    except Exception as e:
        print_error(f"Failed to open editor: {e}")
        return False


def get_commands_for_device_type(device_type: str) -> List[str]:
    """
    Get available commands for a specific device type.

    Args:
        device_type: Device type (e.g., cisco_ios, cisco_xe)

    Returns:
        List of available commands
    """
    commands = []

    # Get commands from database
    if device_type in DEVICE_COMMANDS:
        device_cmds = DEVICE_COMMANDS[device_type]
        commands.extend(device_cmds.get("show", []))

        # Add common administrative commands
        commands.extend(
            [
                "# === Configuration Commands ===",
                "write memory",
                "copy running-config startup-config",
                "# === Troubleshooting Commands ===",
                "ping",
                "traceroute",
                "# === Clear Commands ===",
                "clear counters",
                "clear logging",
            ]
        )

    else:
        # Fallback to generic commands
        commands = [
            "show version",
            "show running-config",
            "show ip interface brief",
            "show interfaces",
            "show inventory",
        ]

    return commands


def edit_commands_with_autocomplete(file_path: str, device_type: str) -> bool:
    """
    Edit commands file with intelligent autocomplete based on device type.

    Args:
        file_path: Path to commands file
        device_type: Device type for command suggestions

    Returns:
        True if successful, False otherwise
    """
    if not PROMPT_TOOLKIT_AVAILABLE:
        print_warning("prompt_toolkit not installed. Using basic editor...")
        return open_file_in_editor(file_path)

    file_path_obj = Path(file_path)

    # Load existing commands
    existing_commands = []
    if file_path_obj.exists():
        try:
            with open(file_path_obj, "r", encoding="utf-8") as f:
                existing_commands = [line.rstrip() for line in f.readlines()]
        except Exception as e:
            print_error(f"Failed to read file: {e}")
            return False

    # Get command suggestions for device type
    command_suggestions = get_commands_for_device_type(device_type)

    print_banner(f"INTERACTIVE COMMAND EDITOR - {device_type.upper()}", "bold green")

    if RICH_AVAILABLE and console:
        console.print(f"\n[cyan]📝 Editing:[/cyan] [yellow]{file_path}[/yellow]")
        console.print(f"[cyan]🔧 Device Type:[/cyan] [yellow]{device_type}[/yellow]")
        console.print("\n[bold]Features:[/bold]")
        console.print("  • Type to see autocomplete suggestions")
        console.print("  • Press [green]Tab[/green] to complete")
        console.print("  • Press [green]↑↓[/green] arrows for history")
        console.print("  • Lines starting with [yellow]#[/yellow] are comments")
        console.print("  • Type [red]done[/red] or press [red]Ctrl+D[/red] to finish")
        console.print("  • Type [red]cancel[/red] to discard changes\n")
    else:
        print(f"\nEditing: {file_path}")
        print(f"Device Type: {device_type}")
        print("\nType 'done' when finished, 'cancel' to discard, or Ctrl+D to save")
        print("Lines starting with # are comments\n")

    # Display existing commands
    if existing_commands:
        print_info("Current commands:")
        for idx, cmd in enumerate(existing_commands, 1):
            if cmd.strip().startswith("#"):
                if RICH_AVAILABLE and console:
                    console.print(f"  [dim]{idx:2d}. {cmd}[/dim]")
                else:
                    print(f"  {idx:2d}. {cmd}")
            else:
                if RICH_AVAILABLE and console:
                    console.print(f"  [green]{idx:2d}.[/green] {cmd}")
                else:
                    print(f"  {idx:2d}. {cmd}")
        print()

    # Create completer with fuzzy matching - clean and simple
    completer = FuzzyCompleter(WordCompleter(command_suggestions, ignore_case=True, sentence=True))

    # Create history
    history = InMemoryHistory()
    for cmd in existing_commands:
        if cmd.strip() and not cmd.strip().startswith("#"):
            history.append_string(cmd)

    # Collect new commands
    new_commands = []
    command_number = len(existing_commands) + 1

    try:
        while True:
            try:
                # Prompt for command with custom styling and enhanced completion menu
                user_input = prompt(
                    HTML(f"<cyan><b>[{command_number}]</b></cyan> <b>&gt;</b> "),
                    completer=completer,
                    history=history,
                    auto_suggest=AutoSuggestFromHistory(),
                    complete_while_typing=True,
                    style=CUSTOM_STYLE,
                    complete_in_thread=True,  # Smooth async completion
                    mouse_support=True,  # Enable mouse in dropdown
                    refresh_interval=0.5,  # Smooth refresh
                ).strip()

                if not user_input:
                    continue

                if user_input.lower() == "done":
                    break
                elif user_input.lower() == "cancel":
                    print_warning("Changes discarded")
                    return False
                else:
                    new_commands.append(user_input)
                    if RICH_AVAILABLE and console:
                        console.print(f"  ✅ Added: [cyan]{user_input}[/cyan]")
                    else:
                        print(f"  ✅ Added: {user_input}")
                    command_number += 1

            except EOFError:
                # Ctrl+D pressed
                break

    except KeyboardInterrupt:
        print_warning("\nEdit cancelled")
        return False

    # Combine existing and new commands
    all_commands = existing_commands + new_commands

    if not all_commands:
        print_warning("No commands to save")
        return False

    # Save to file
    try:
        with open(file_path_obj, "w", encoding="utf-8") as f:
            for cmd in all_commands:
                f.write(cmd + "\n")

        print_success(f"Saved {len(all_commands)} commands to {file_path}")

        # Show summary
        non_comment_commands = [
            c for c in all_commands if c.strip() and not c.strip().startswith("#")
        ]
        comment_lines = [c for c in all_commands if c.strip().startswith("#")]

        if RICH_AVAILABLE and console:
            from rich.table import Table

            table = Table(title="📊 Summary", show_header=False, border_style="cyan")
            table.add_column("Type", style="cyan")
            table.add_column("Count", style="green")
            table.add_row("⚡ Commands", str(len(non_comment_commands)))
            table.add_row("💬 Comments", str(len(comment_lines)))
            table.add_row("📄 Total Lines", str(len(all_commands)))
            console.print(table)
        else:
            print("\n📊 Summary:")
            print(f"  ⚡ Commands: {len(non_comment_commands)}")
            print(f"  💬 Comments: {len(comment_lines)}")
            print(f"  📄 Total Lines: {len(all_commands)}")

        return True

    except Exception as e:
        print_error(f"Failed to save file: {e}")
        return False


def list_output_files(base_name: str = "output_") -> List[str]:
    """
    List all output files matching the pattern.

    Args:
        base_name: Base name pattern to search for

    Returns:
        List of matching file paths
    """
    cwd = Path.cwd()
    extensions = [".csv", ".json", ".html", ".md", ".xlsx"]

    files = []
    for ext in extensions:
        pattern = f"{base_name}*{ext}"
        files.extend(cwd.glob(pattern))

    return sorted([str(f) for f in files], reverse=True)


def view_output_file(file_path: str) -> None:
    """
    View an output file in the appropriate viewer.

    Args:
        file_path: Path to the output file
    """
    import platform
    import subprocess  # nosec B404 - subprocess needed for cross-platform file opening
    import webbrowser

    file_path_obj = Path(file_path)

    if not file_path_obj.exists():
        print_error(f"File not found: {file_path}")
        return

    try:
        suffix = file_path_obj.suffix.lower()
        system = platform.system()

        # HTML files - open in browser
        if suffix == ".html":
            webbrowser.open(f"file://{file_path_obj.absolute()}")
            print_success(f"Opened {file_path} in browser")

        # Excel files - open with default app
        elif suffix == ".xlsx":
            if system == "Windows":
                # Use os.startfile on Windows to avoid shell=True security issue
                import os

                os.startfile(str(file_path_obj))  # nosec B606 # pylint: disable=no-member
            elif system == "Darwin":
                subprocess.run(["open", str(file_path_obj)], check=False)  # nosec B603 B607
            else:
                subprocess.run(["xdg-open", str(file_path_obj)], check=False)  # nosec B603 B607
            print_success(f"Opened {file_path} in default application")

        # Text-based files - display in terminal with Rich or cat
        elif suffix in [".csv", ".json", ".md"]:
            if RICH_AVAILABLE and console and suffix in [".json", ".md"]:
                # Use Rich syntax highlighting for JSON and Markdown
                from rich.syntax import Syntax

                with open(file_path_obj, "r", encoding="utf-8") as f:
                    content = f.read()

                if suffix == ".json":
                    syntax = Syntax(content, "json", theme="monokai", line_numbers=True)
                elif suffix == ".md":
                    syntax = Syntax(content, "markdown", theme="monokai", line_numbers=True)

                console.print(syntax)
                print_success(f"\nDisplayed {file_path}")

            else:
                # Fallback to simple display
                print_banner(f"Contents of {file_path}")
                with open(file_path_obj, "r", encoding="utf-8") as f:
                    content = f.read()
                    # Limit output to first 100 lines for CSV
                    if suffix == ".csv":
                        lines = content.split("\n")
                        if len(lines) > 100:
                            print("\n".join(lines[:100]))
                            print(f"\n... ({len(lines) - 100} more lines)")
                        else:
                            print(content)
                    else:
                        print(content)

        else:
            print_warning(f"Unknown file type: {suffix}")

    except Exception as e:
        print_error(f"Failed to view file: {e}")


def get_or_create_file(default_filename: str, file_description: str) -> Optional[str]:
    """
    Check for file in current directory, or prompt user to create/select one.

    Args:
        default_filename: Default filename to look for (e.g., "devices.csv")
        file_description: Description of the file type (e.g., "devices CSV")

    Returns:
        Path to the file, or None if user cancels
    """
    current_dir = Path.cwd()
    default_path = current_dir / default_filename

    # Check if file exists in current directory
    if default_path.exists():
        print_success(f"Found {default_filename} in current directory")
        print_info(f"📁 Location: {default_path}")

        # Ask for confirmation
        confirm = input("\nUse this file? (yes/no) [yes]: ").strip().lower()
        if confirm in ["", "yes", "y"]:
            return str(default_path)

    # File not found or user declined
    print_warning(f"{default_filename} not found in current directory")
    print_info(f"📂 Current directory: {current_dir}")
    print()

    # Offer options
    options = [
        ("create", f"Create new {default_filename} in current directory"),
        ("browse", "Enter path to existing file"),
        ("cancel", "Cancel and go back"),
    ]

    choice = select_from_menu("What would you like to do?", options, "bold yellow")

    if choice == "create":
        print_success(f"Creating new {default_filename} in {current_dir}")
        return str(default_path)

    elif choice == "browse":
        print_info("Enter the full path to your file")
        print_info("Tip: You can drag and drop the file into this window")

        file_path = input(f"\nEnter path to {file_description} file: ").strip()

        # Remove quotes if user drag-and-dropped
        file_path = file_path.strip('"').strip("'")

        if file_path:
            file_obj = Path(file_path)
            if file_obj.exists():
                print_success(f"Found file: {file_obj}")
                return str(file_obj)
            else:
                print_warning(f"File not found: {file_path}")
                create = input("Create this file? (yes/no) [no]: ").strip().lower()
                if create in ["yes", "y"]:
                    return file_path
                return None
        return None

    else:  # cancel or None
        return None


def show_file_manager_menu() -> None:
    """Display file manager submenu for editing and viewing files."""
    while True:
        # Use arrow-key navigable menu
        menu_options = [
            ("1", "Edit devices.csv"),
            ("2", "Edit commands.txt (with autocomplete)"),
            ("3", "View output files"),
            ("4", "Create sample devices.csv"),
            ("5", "Create sample commands.txt"),
            ("6", "Back to Main Menu"),
        ]

        choice = select_from_menu("FILE MANAGER", menu_options, "bold yellow")

        # Handle Escape key - go back to main menu
        if choice is None or choice == "6":
            return

        if choice == "1":
            # Auto-detect or prompt for devices.csv
            devices_file = get_or_create_file("devices.csv", "devices CSV")

            if devices_file:
                print_info(f"Opening {devices_file} in editor...")
                if open_file_in_editor(devices_file):
                    print_success("Editor closed")
            else:
                print_info("Operation cancelled")

        elif choice == "2":
            # Auto-detect or prompt for commands.txt
            commands_file = get_or_create_file("commands.txt", "commands")

            if commands_file:
                # Check if we can use the smart autocomplete editor
                if PROMPT_TOOLKIT_AVAILABLE:
                    # Get device type from config for autocomplete suggestions
                    config = load_config()
                    device_type = config.get("default_device_type", "cisco_ios")

                    print_info(f"Opening {commands_file} in smart editor with autocomplete...")
                    print_info(f"Device type: {device_type}")
                    print_info(
                        "Tip: Press Tab for suggestions, type 'done' when finished, "
                        "'cancel' to discard changes"
                    )

                    if edit_commands_with_autocomplete(commands_file, device_type):
                        print_success("Commands file updated successfully")
                    else:
                        print_warning("No changes made")
                else:
                    # Fallback to basic editor
                    print_warning("prompt_toolkit not available - using basic editor")
                    print_info(
                        "Install prompt_toolkit for autocomplete: pip install prompt_toolkit"
                    )
                    print_info(f"Opening {commands_file} in editor...")
                    if open_file_in_editor(commands_file):
                        print_success("Editor closed")
            else:
                print_info("Operation cancelled")

        elif choice == "3":
            output_files = list_output_files()

            if not output_files:
                print_warning("No output files found")
                continue

            print_banner("Available Output Files")
            if RICH_AVAILABLE and console:
                table = Table(show_header=True)
                table.add_column("#", style="cyan", width=4)
                table.add_column("File", style="green")
                table.add_column("Type", style="yellow")

                for idx, file in enumerate(output_files, 1):
                    file_obj = Path(file)
                    file_type = file_obj.suffix.upper().replace(".", "")
                    table.add_row(str(idx), file_obj.name, file_type)

                console.print(table)
            else:
                for idx, file in enumerate(output_files, 1):
                    print(f"{idx}. {Path(file).name}")

            try:
                file_choice = input("\nEnter file number to view (or Enter to cancel): ").strip()
                if file_choice:
                    file_idx = int(file_choice) - 1
                    if 0 <= file_idx < len(output_files):
                        view_output_file(output_files[file_idx])
                        input("\nPress Enter to continue...")
                    else:
                        print_error("Invalid file number")
            except ValueError:
                print_error("Invalid input")

        elif choice == "4":
            # Create sample devices.csv
            sample_devices = """hostname,ip_address,device_type
router1,192.168.1.1,cisco_ios
switch1,192.168.1.2,cisco_ios
router2,10.0.0.1,cisco_xe
"""
            devices_file = "devices.csv"
            if Path(devices_file).exists():
                confirm = (
                    input(f"{devices_file} already exists. Overwrite? (yes/no) [no]: ")
                    .strip()
                    .lower()
                )
                if confirm not in ["yes", "y"]:
                    continue

            try:
                with open(devices_file, "w", encoding="utf-8") as f:
                    f.write(sample_devices)
                print_success(f"Created sample {devices_file}")
            except Exception as e:
                print_error(f"Failed to create file: {e}")

        elif choice == "5":
            # Create sample commands.txt
            sample_commands = """# Sample network commands
show version
show ip interface brief
show running-config | include hostname
show inventory
"""
            commands_file = "commands.txt"
            if Path(commands_file).exists():
                confirm = (
                    input(f"{commands_file} already exists. Overwrite? (yes/no) [no]: ")
                    .strip()
                    .lower()
                )
                if confirm not in ["yes", "y"]:
                    continue

            try:
                with open(commands_file, "w", encoding="utf-8") as f:
                    f.write(sample_commands)
                print_success(f"Created sample {commands_file}")
            except Exception as e:
                print_error(f"Failed to create file: {e}")


def load_config() -> Dict:
    """
    Load configuration from file.

    Returns:
        Configuration dictionary with default values if file doesn't exist
    """
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config = json.load(f)
                # Merge with defaults to ensure all keys exist
                return {**DEFAULT_CONFIG, **config}
        except (json.JSONDecodeError, IOError) as e:
            logger.warning("Failed to load config file: %s. Using defaults.", e)

    return DEFAULT_CONFIG.copy()


def save_config(config: Dict) -> None:
    """
    Save configuration to file.

    Args:
        config: Configuration dictionary to save
    """
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        logger.info("Configuration saved to %s", CONFIG_FILE)
    except IOError as e:
        logger.error("Failed to save config file: %s", e)


def show_settings_menu(config: Dict) -> Dict:
    """
    Display interactive settings menu and allow user to modify configuration.

    Args:
        config: Current configuration dictionary

    Returns:
        Updated configuration dictionary
    """
    while True:
        # Use arrow-key navigable menu with current values
        menu_options = [
            ("1", f"Default Device Type: {config['default_device_type']}"),
            ("2", f"Strip Whitespace: {config['strip_whitespace']}"),
            ("3", f"Max Workers: {config['max_workers']}"),
            ("4", f"Connection Timeout: {config['connection_timeout']}s"),
            ("5", f"Command Timeout: {config['command_timeout']}s"),
            ("6", f"Enable Session Logging: {config.get('enable_session_logging', False)}"),
            ("7", f"Enable Mode: {config.get('enable_mode', False)}"),
            ("8", f"Retry on Failure: {config.get('retry_on_failure', True)}"),
            ("9", "Save Settings"),
            ("10", "Reset to Defaults"),
            ("11", "Back to Main Menu"),
        ]

        choice = select_from_menu("SETTINGS MENU", menu_options)

        # Handle Escape key - go back without saving
        if choice is None or choice == "11":
            return config

        if choice == "1":
            prompt = f"Enter default device type [{config['default_device_type']}]: "
            new_value = input(prompt).strip()
            if new_value:
                config["default_device_type"] = new_value
                print_success(f"Default device type set to: {new_value}")

        elif choice == "2":
            current = "yes" if config["strip_whitespace"] else "no"
            prompt = f"Strip whitespace? (yes/no) [{current}]: "
            new_value = input(prompt).strip().lower()
            if new_value == "":
                # Keep current value on Enter
                pass
            elif new_value in ["yes", "y", "true", "1"]:
                config["strip_whitespace"] = True
                print_success("Whitespace stripping enabled")
            elif new_value in ["no", "n", "false", "0"]:
                config["strip_whitespace"] = False
                print_success("Whitespace stripping disabled")

        elif choice == "3":
            new_value = get_positive_int_input(
                "Enter max workers", config["max_workers"], MIN_WORKERS, MAX_WORKERS
            )
            if new_value:
                config["max_workers"] = new_value
                print_success(f"Max workers set to: {new_value}")

        elif choice == "4":
            new_value = get_positive_int_input(
                "Enter connection timeout (seconds)", config["connection_timeout"]
            )
            if new_value:
                config["connection_timeout"] = new_value
                print_success(f"Connection timeout set to: {new_value} seconds")

        elif choice == "5":
            new_value = get_positive_int_input(
                "Enter command timeout (seconds)", config["command_timeout"]
            )
            if new_value:
                config["command_timeout"] = new_value
                print_success(f"Command timeout set to: {new_value} seconds")

        elif choice == "6":
            current = "yes" if config.get("enable_session_logging", False) else "no"
            prompt = f"Enable session logging? (yes/no) [{current}]: "
            new_value = input(prompt).strip().lower()
            if new_value == "":
                # Keep current value on Enter
                pass
            elif new_value in ["yes", "y", "true", "1"]:
                config["enable_session_logging"] = True
                print_success("Session logging enabled")
                print_warning("Session logs may contain sensitive information!")
            elif new_value in ["no", "n", "false", "0"]:
                config["enable_session_logging"] = False
                print_success("Session logging disabled")

        elif choice == "7":
            current = "yes" if config.get("enable_mode", False) else "no"
            prompt = f"Enter enable mode on connect? (yes/no) [{current}]: "
            new_value = input(prompt).strip().lower()
            if new_value == "":
                # Keep current value on Enter
                pass
            elif new_value in ["yes", "y", "true", "1"]:
                config["enable_mode"] = True
                print_success("Enable mode enabled")
            elif new_value in ["no", "n", "false", "0"]:
                config["enable_mode"] = False
                print_success("Enable mode disabled")

        elif choice == "8":
            current = "yes" if config.get("retry_on_failure", True) else "no"
            prompt = f"Retry on connection failure? (yes/no) [{current}]: "
            new_value = input(prompt).strip().lower()
            if new_value == "":
                # Keep current value on Enter
                pass
            elif new_value in ["yes", "y", "true", "1"]:
                config["retry_on_failure"] = True
                print_success("Retry on failure enabled")
            elif new_value in ["no", "n", "false", "0"]:
                config["retry_on_failure"] = False
                print_success("Retry on failure disabled")

        elif choice == "9":
            save_config(config)
            print_success("Settings saved successfully")

        elif choice == "10":
            confirm = input("Reset all settings to defaults? (yes/no) [no]: ").strip().lower()
            if confirm in ["yes", "y"]:
                config = DEFAULT_CONFIG.copy()
                print_success("Settings reset to defaults")


def load_devices(
    devices_file: str, default_device_type: Optional[str] = None
) -> List[Dict[str, str]]:
    """
    Load device information from a CSV file.

    Expected CSV format:
    hostname,ip_address,device_type,ssh_config_file,use_keys,key_file
    router1,192.168.1.1,cisco_ios,,,
    switch1,192.168.1.2,cisco_ios,~/.ssh/config,true,~/.ssh/id_rsa

    Required fields: hostname, ip_address
    Optional fields: device_type, ssh_config_file, use_keys, key_file

    If device_type is not provided in CSV and default_device_type is set,
    the default will be used.

    Args:
        devices_file: Path to the devices CSV file
        default_device_type: Default device type to use if not specified

    Returns:
        List of device dictionaries

    Raises:
        FileNotFoundError: If the devices file doesn't exist
        ValueError: If the CSV format is invalid
    """
    devices = []

    if not Path(devices_file).exists():
        raise FileNotFoundError(f"Devices file not found: {devices_file}")

    try:
        with open(devices_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            required_fields = {"hostname", "ip_address"}
            optional_fields = {"device_type", "ssh_config_file", "use_keys", "key_file"}

            for row_num, row in enumerate(reader, start=2):
                # Validate required fields
                if not required_fields.issubset(row.keys()):
                    msg = "CSV must contain columns: "
                    msg += f"{', '.join(required_fields)}"
                    raise ValueError(msg)

                normalized_row = {}
                missing_values = []

                for field in required_fields:
                    value = row.get(field)
                    if value is None:
                        missing_values.append(field)
                        continue

                    trimmed = value.strip()
                    if not trimmed:
                        missing_values.append(field)
                        continue

                    normalized_row[field] = trimmed

                if missing_values:
                    fields = ", ".join(sorted(missing_values))
                    msg = "Row {} has missing values for: {}".format(row_num, fields)
                    raise ValueError(msg)

                # Process optional fields
                for field in optional_fields:
                    value = row.get(field, "").strip()
                    if value:
                        normalized_row[field] = value

                # Apply default device_type if not specified
                if "device_type" not in normalized_row and default_device_type:
                    normalized_row["device_type"] = default_device_type

                # Verify device_type is set
                if "device_type" not in normalized_row:
                    msg = f"Row {row_num}: device_type not specified "
                    msg += "and no default provided"
                    raise ValueError(msg)

                devices.append(normalized_row)

        if not devices:
            raise ValueError("No devices found in the file")

        logger.info("Loaded %d device(s) from %s", len(devices), devices_file)
        return devices

    except csv.Error as e:
        raise ValueError(f"Error parsing CSV file: {e}") from e


def load_commands(commands_file: str) -> List[str]:
    """
    Load commands from a text file (one command per line).

    Args:
        commands_file: Path to the commands file

    Returns:
        List of commands to execute

    Raises:
        FileNotFoundError: If the commands file doesn't exist
    """
    if not Path(commands_file).exists():
        raise FileNotFoundError(f"Commands file not found: {commands_file}")

    with open(commands_file, "r", encoding="utf-8") as f:
        commands = []
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            commands.append(stripped)

    if not commands:
        raise ValueError("No commands found in the file")

    logger.info("Loaded %d command(s) from %s", len(commands), commands_file)
    return commands


def connect_and_execute(
    device: Dict[str, str],
    commands: List[str],
    username: str,
    password: str,
    global_ssh_config: str = None,
    strip_whitespace: bool = True,
    connection_timeout: int = 30,
    command_timeout: int = 60,
    enable_session_logging: bool = False,
    enable_mode: bool = False,
    enable_password: str = None,
) -> List[Dict[str, str]]:
    """
    Connect to a device and execute commands.

    Args:
        device: Device information dictionary
        commands: List of commands to execute
        username: SSH username
        password: SSH password
        global_ssh_config: Global SSH config file path (optional)
        strip_whitespace: Strip extra whitespace from output
        connection_timeout: Connection timeout in seconds
        command_timeout: Command execution timeout in seconds
        enable_session_logging: Enable session logging (default: False)
        enable_mode: Enter enable mode after connecting (default: False)
        enable_password: Enable mode password (optional)

    Returns:
        List of result dictionaries containing command outputs
    """
    results = []
    hostname = device["hostname"]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    device_params = {
        "device_type": device["device_type"],
        "host": device["ip_address"],
        "username": username,
        "password": password,
        "timeout": connection_timeout,
    }

    # Add session logging only if enabled
    if enable_session_logging:
        device_params["session_log"] = f"session_{hostname}_{timestamp}.log"
        logger.debug("Session logging enabled for %s", hostname)

    # Add SSH config file support (device-specific or global) with path expansion
    ssh_config_file = device.get("ssh_config_file")
    if not ssh_config_file:
        ssh_config_file = global_ssh_config
    if ssh_config_file:
        expanded_config = expand_path(ssh_config_file)
        if expanded_config.exists():
            device_params["ssh_config_file"] = str(expanded_config)
            logger.debug("Using SSH config: %s", expanded_config)
        else:
            logger.warning("SSH config file not found: %s", expanded_config)

    # Add SSH key authentication support with validation
    if device.get("use_keys"):
        use_keys_value = device["use_keys"].lower()
        if use_keys_value in TRUTHY_VALUES:
            device_params["use_keys"] = True
            if device.get("key_file"):
                key_path = expand_path(device["key_file"])
                if key_path.exists():
                    device_params["key_file"] = str(key_path)
                    logger.debug("Using SSH key: %s", key_path)
                else:
                    logger.warning("SSH key file not found: %s", key_path)

    try:
        logger.info("Connecting to %s (%s)...", hostname, device["ip_address"])
        connection = ConnectHandler(**device_params)

        logger.info("Successfully connected to %s", hostname)

        # Enter enable mode if requested
        if enable_mode:
            try:
                if enable_password:
                    connection.enable(secret=enable_password)
                else:
                    connection.enable()
                logger.info("Entered enable mode on %s", hostname)
            except Exception as enable_error:
                logger.warning("Failed to enter enable mode on %s: %s", hostname, enable_error)

        # Execute each command
        for command in commands:
            logger.info("Executing '%s' on %s", command, hostname)
            try:
                output = connection.send_command(command, read_timeout=command_timeout)

                # Strip extra whitespace if enabled
                if strip_whitespace:
                    # Remove trailing whitespace from each line and
                    # leading/trailing empty lines
                    lines = output.splitlines()
                    stripped_lines = [line.rstrip() for line in lines]
                    # Remove leading and trailing empty lines
                    while stripped_lines and not stripped_lines[0]:
                        stripped_lines.pop(0)
                    while stripped_lines and not stripped_lines[-1]:
                        stripped_lines.pop()
                    output = "\n".join(stripped_lines)

                results.append(
                    {
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "hostname": hostname,
                        "ip_address": device["ip_address"],
                        "command": command,
                        "output": output,
                        "status": "success",
                    }
                )
                logger.info("Command '%s' executed successfully on %s", command, hostname)
            except Exception as cmd_error:
                error_msg = f"Error executing command: {str(cmd_error)}"
                logger.error("%s on %s", error_msg, hostname)
                results.append(
                    {
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "hostname": hostname,
                        "ip_address": device["ip_address"],
                        "command": command,
                        "output": error_msg,
                        "status": "failed",
                    }
                )

        connection.disconnect()
        logger.info("Disconnected from %s", hostname)

    except NetmikoTimeoutException:
        error_msg = "Connection timeout - device unreachable"
        logger.error("%s: %s", hostname, error_msg)
        for command in commands:
            results.append(
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "hostname": hostname,
                    "ip_address": device["ip_address"],
                    "command": command,
                    "output": error_msg,
                    "status": "failed",
                }
            )

    except NetmikoAuthenticationException:
        error_msg = "Authentication failed - check credentials"
        logger.error("%s: %s", hostname, error_msg)
        for command in commands:
            results.append(
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "hostname": hostname,
                    "ip_address": device["ip_address"],
                    "command": command,
                    "output": error_msg,
                    "status": "failed",
                }
            )

    except Exception as conn_error:
        error_msg = f"Unexpected error: {str(conn_error)}"
        logger.error("%s: %s", hostname, error_msg)
        for command in commands:
            results.append(
                {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "hostname": hostname,
                    "ip_address": device["ip_address"],
                    "command": command,
                    "output": error_msg,
                    "status": "failed",
                }
            )

    return results


def connect_with_retry(
    device: Dict[str, str],
    commands: List[str],
    username: str,
    password: str,
    global_ssh_config: str,
    strip_whitespace: bool,
    connection_timeout: int,
    command_timeout: int,
    enable_session_logging: bool,
    enable_mode: bool,
    enable_password: str,
    retry_enabled: bool,
) -> List[Dict[str, str]]:
    """
    Connect to device with optional retry logic.

    Args:
        device: Device information dictionary
        commands: List of commands to execute
        username: SSH username
        password: SSH password
        global_ssh_config: Global SSH config file path
        strip_whitespace: Whether to strip whitespace from output
        connection_timeout: Connection timeout in seconds
        command_timeout: Command timeout in seconds
        enable_session_logging: Enable session logging
        enable_mode: Enter enable mode
        enable_password: Enable mode password
        retry_enabled: Enable retry on failure

    Returns:
        List of result dictionaries
    """
    if retry_enabled and TENACITY_AVAILABLE:
        # Create a retry decorator
        retry_decorator = retry(
            stop=stop_after_attempt(MAX_RETRY_ATTEMPTS),
            wait=wait_exponential(min=RETRY_MIN_WAIT, max=RETRY_MAX_WAIT),
            retry=retry_if_exception_type(
                (NetmikoTimeoutException, NetmikoAuthenticationException)
            ),
            reraise=True,
        )

        # Wrap the function with retry
        retryable_connect = retry_decorator(connect_and_execute)

        try:
            return retryable_connect(
                device,
                commands,
                username,
                password,
                global_ssh_config,
                strip_whitespace,
                connection_timeout,
                command_timeout,
                enable_session_logging,
                enable_mode,
                enable_password,
            )
        except Exception:
            # If all retries fail, call without retry to get proper error handling
            return connect_and_execute(
                device,
                commands,
                username,
                password,
                global_ssh_config,
                strip_whitespace,
                connection_timeout,
                command_timeout,
                enable_session_logging,
                enable_mode,
                enable_password,
            )
    else:
        # No retry - direct connection
        return connect_and_execute(
            device,
            commands,
            username,
            password,
            global_ssh_config,
            strip_whitespace,
            connection_timeout,
            command_timeout,
            enable_session_logging,
            enable_mode,
            enable_password,
        )


def process_devices_parallel(
    devices: List[Dict[str, str]],
    commands: List[str],
    username: str,
    password: str,
    global_ssh_config: Optional[str],
    strip_whitespace: bool,
    connection_timeout: int,
    command_timeout: int,
    max_workers: int = 5,
    enable_session_logging: bool = False,
    enable_mode: bool = False,
    enable_password: Optional[str] = None,
    retry_on_failure: bool = True,
    show_progress: bool = True,
) -> List[Dict[str, str]]:
    """
    Process multiple devices in parallel using ThreadPoolExecutor.

    Args:
        devices: List of device dictionaries
        commands: List of commands to execute
        username: SSH username
        password: SSH password
        global_ssh_config: Global SSH config file path
        strip_whitespace: Whether to strip whitespace from output
        connection_timeout: Connection timeout in seconds
        command_timeout: Command timeout in seconds
        max_workers: Maximum number of concurrent workers
        enable_session_logging: Enable session logging
        enable_mode: Enter enable mode
        enable_password: Enable mode password
        retry_on_failure: Retry on connection failures
        show_progress: Show progress bar

    Returns:
        List of all results from all devices
    """
    all_results = []

    # Use rich progress bar if available and requested
    if show_progress and RICH_AVAILABLE and console:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("[cyan]Processing devices...", total=len(devices))

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all device processing tasks
                future_to_device = {
                    executor.submit(
                        connect_with_retry,
                        device,
                        commands,
                        username,
                        password,
                        global_ssh_config,
                        strip_whitespace,
                        connection_timeout,
                        command_timeout,
                        enable_session_logging,
                        enable_mode,
                        enable_password,
                        retry_on_failure,
                    ): device
                    for device in devices
                }

                # Collect results as they complete
                for future in as_completed(future_to_device):
                    device = future_to_device[future]
                    try:
                        results = future.result()
                        all_results.extend(results)
                        progress.update(
                            task, advance=1, description=f"[green]Processed {device['hostname']}"
                        )
                    except Exception as exc:
                        logger.error(
                            "Device %s generated an exception: %s", device["hostname"], exc
                        )
                        progress.update(
                            task, advance=1, description=f"[red]Failed {device['hostname']}"
                        )
                        # Add error results for all commands
                        for command in commands:
                            all_results.append(
                                {
                                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "hostname": device["hostname"],
                                    "ip_address": device["ip_address"],
                                    "command": command,
                                    "output": f"Exception during processing: {str(exc)}",
                                    "status": "failed",
                                }
                            )

    elif show_progress and TQDM_AVAILABLE:
        # Fallback to tqdm if rich not available
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_device = {
                executor.submit(
                    connect_with_retry,
                    device,
                    commands,
                    username,
                    password,
                    global_ssh_config,
                    strip_whitespace,
                    connection_timeout,
                    command_timeout,
                    enable_session_logging,
                    enable_mode,
                    enable_password,
                    retry_on_failure,
                ): device
                for device in devices
            }

            # Use tqdm progress bar
            for future in tqdm(
                as_completed(future_to_device), total=len(devices), desc="Processing devices"
            ):
                device = future_to_device[future]
                try:
                    results = future.result()
                    all_results.extend(results)
                except Exception as exc:
                    logger.error("Device %s generated an exception: %s", device["hostname"], exc)
                    for command in commands:
                        all_results.append(
                            {
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "hostname": device["hostname"],
                                "ip_address": device["ip_address"],
                                "command": command,
                                "output": f"Exception during processing: {str(exc)}",
                                "status": "failed",
                            }
                        )
    else:
        # No progress bar
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_device = {
                executor.submit(
                    connect_with_retry,
                    device,
                    commands,
                    username,
                    password,
                    global_ssh_config,
                    strip_whitespace,
                    connection_timeout,
                    command_timeout,
                    enable_session_logging,
                    enable_mode,
                    enable_password,
                    retry_on_failure,
                ): device
                for device in devices
            }

            for future in as_completed(future_to_device):
                device = future_to_device[future]
                try:
                    results = future.result()
                    all_results.extend(results)
                except Exception as exc:
                    logger.error("Device %s generated an exception: %s", device["hostname"], exc)
                    for command in commands:
                        all_results.append(
                            {
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "hostname": device["hostname"],
                                "ip_address": device["ip_address"],
                                "command": command,
                                "output": f"Exception during processing: {str(exc)}",
                                "status": "failed",
                            }
                        )

    return all_results


def save_to_csv(results: List[Dict[str, str]], output_file: str) -> None:
    """
    Save command outputs to a CSV file.

    Args:
        results: List of result dictionaries
        output_file: Path to the output CSV file
    """
    if not results:
        logger.warning("No results to save")
        return

    fieldnames = ["timestamp", "hostname", "ip_address", "command", "output", "status"]

    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    logger.info("Results saved to %s", output_file)


def save_to_json(results: List[Dict[str, str]], output_file: str) -> None:
    """
    Save command outputs to a JSON file with beautiful formatting.

    Args:
        results: List of result dictionaries
        output_file: Path to the output JSON file
    """
    if not results:
        logger.warning("No results to save")
        return

    # Group results by device for better structure
    devices_data = {}
    for result in results:
        hostname = result["hostname"]
        if hostname not in devices_data:
            devices_data[hostname] = {
                "hostname": hostname,
                "ip_address": result["ip_address"],
                "commands": [],
            }

        devices_data[hostname]["commands"].append(
            {
                "timestamp": result["timestamp"],
                "command": result["command"],
                "output": result["output"],
                "status": result["status"],
            }
        )

    output_data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_devices": len(devices_data),
        "total_commands": len(results),
        "devices": list(devices_data.values()),
    }

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    logger.info("JSON results saved to %s", output_file)


def save_to_markdown(results: List[Dict[str, str]], output_file: str) -> None:
    """
    Save command outputs to a Markdown file with beautiful formatting.

    Args:
        results: List of result dictionaries
        output_file: Path to the output Markdown file
    """
    if not results:
        logger.warning("No results to save")
        return

    # Group by device
    devices_data = {}
    for result in results:
        hostname = result["hostname"]
        if hostname not in devices_data:
            devices_data[hostname] = {"ip_address": result["ip_address"], "commands": []}
        devices_data[hostname]["commands"].append(result)

    # Calculate statistics
    total_commands = len(results)
    successful = sum(1 for r in results if r["status"] == "success")
    failed = total_commands - successful

    with open(output_file, "w", encoding="utf-8") as f:
        # Header
        f.write("# Network Device Command Collection Report\n\n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # Summary
        f.write("## Summary\n\n")
        f.write(f"- **Total Devices:** {len(devices_data)}\n")
        f.write(f"- **Total Commands:** {total_commands}\n")
        f.write(f"- **Successful:** {successful} ✓\n")
        f.write(f"- **Failed:** {failed} ✗\n\n")

        # Device Results
        f.write("## Device Results\n\n")

        for hostname, data in devices_data.items():
            f.write(f"### {hostname}\n\n")
            f.write(f"**IP Address:** `{data['ip_address']}`\n\n")

            for cmd_result in data["commands"]:
                status_emoji = "✓" if cmd_result["status"] == "success" else "✗"
                f.write(f"#### {status_emoji} `{cmd_result['command']}`\n\n")
                f.write(f"**Timestamp:** {cmd_result['timestamp']}  \n")
                f.write(f"**Status:** {cmd_result['status']}\n\n")

                if cmd_result["status"] == "success":
                    f.write("**Output:**\n\n")
                    f.write("```\n")
                    f.write(cmd_result["output"])
                    f.write("\n```\n\n")
                else:
                    f.write(f"**Error:** {cmd_result['output']}\n\n")

            f.write("---\n\n")

    logger.info("Markdown report saved to %s", output_file)


def save_to_html(results: List[Dict[str, str]], output_file: str) -> None:  # type: ignore
    """
    Save command outputs to an HTML file with beautiful Bootstrap styling.

    Args:
        results: List of result dictionaries
        output_file: Path to the output HTML file
    """
    if not results:
        logger.warning("No results to save")
        return

    # Group by device
    devices_data = {}
    for result in results:
        hostname = result["hostname"]
        if hostname not in devices_data:
            devices_data[hostname] = {"ip_address": result["ip_address"], "commands": []}
        devices_data[hostname]["commands"].append(result)

    # Calculate statistics
    total_commands = len(results)
    successful = sum(1 for r in results if r["status"] == "success")
    failed = total_commands - successful
    success_rate = (successful / total_commands * 100) if total_commands > 0 else 0

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Network Device Command Collection Report</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css"
          rel="stylesheet">
    <link rel="stylesheet"
          href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css">
    <style>
        body {{
            background-color: #f8f9fa;
            padding: 20px;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .stat-card {{
            border-radius: 10px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .device-card {{
            background: white;
            border-radius: 10px;
            padding: 25px;
            margin-bottom: 25px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        .command-output {{
            background-color: #2d2d2d;
            color: #f8f8f2;
            padding: 15px;
            border-radius: 5px;
            font-family: 'Courier New', monospace;
            font-size: 14px;
            overflow-x: auto;
            white-space: pre-wrap;
        }}
        .badge-success-custom {{
            background-color: #10b981;
        }}
        .badge-danger-custom {{
            background-color: #ef4444;
        }}
        .progress-bar-custom {{
            background: linear-gradient(90deg, #10b981 0%, #059669 100%);
        }}
    </style>
</head>
<body>
    <div class="container-fluid">
        <!-- Header -->
        <div class="header">
            <h1><i class="bi bi-router"></i> Network Device Command
                Collection Report</h1>
            <p class="mb-0"><i class="bi bi-calendar"></i> Generated:
               {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>

        <!-- Statistics -->
        <div class="row mb-4">
            <div class="col-md-3">
                <div class="stat-card bg-primary text-white">
                    <h5><i class="bi bi-hdd-network"></i> Total Devices</h5>
                    <h2>{len(devices_data)}</h2>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card bg-info text-white">
                    <h5><i class="bi bi-terminal"></i> Total Commands</h5>
                    <h2>{total_commands}</h2>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card bg-success text-white">
                    <h5><i class="bi bi-check-circle"></i> Successful</h5>
                    <h2>{successful}</h2>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card bg-danger text-white">
                    <h5><i class="bi bi-x-circle"></i> Failed</h5>
                    <h2>{failed}</h2>
                </div>
            </div>
        </div>

        <!-- Success Rate -->
        <div class="card mb-4">
            <div class="card-body">
                <h5>Success Rate</h5>
                <div class="progress" style="height: 30px;">
                    <div class="progress-bar progress-bar-custom" role="progressbar"
                         style="width: {success_rate}%;" aria-valuenow="{success_rate}"
                         aria-valuemin="0" aria-valuemax="100">
                        {success_rate:.1f}%
                    </div>
                </div>
            </div>
        </div>

        <!-- Device Results -->
        <h2 class="mb-4">Device Results</h2>
"""

    for hostname, data in devices_data.items():
        device_success = sum(1 for c in data["commands"] if c["status"] == "success")
        device_total = len(data["commands"])

        html_content += f"""
        <div class="device-card">
            <h3><i class="bi bi-router-fill text-primary"></i> {hostname}</h3>
            <p><strong>IP Address:</strong> <code>{data['ip_address']}</code></p>
            <p><strong>Commands:</strong> {device_success}/{device_total} successful</p>

            <div class="accordion" id="accordion-{hostname}">
"""

        for idx, cmd_result in enumerate(data["commands"]):
            status_class = "success" if cmd_result["status"] == "success" else "danger"
            status_icon = "check-circle" if cmd_result["status"] == "success" else "x-circle"

            html_content += f"""
                <div class="accordion-item">
                    <h2 class="accordion-header">
                        <button class="accordion-button collapsed" type="button"
                                data-bs-toggle="collapse"
                                data-bs-target="#collapse-{hostname}-{idx}">
                            <i class="bi bi-{status_icon} text-{status_class} me-2"></i>
                            <code>{cmd_result['command']}</code>
                            <span class="badge badge-{status_class}-custom ms-2">
                                {cmd_result['status']}</span>
                        </button>
                    </h2>
                    <div id="collapse-{hostname}-{idx}" class="accordion-collapse collapse"
                         data-bs-parent="#accordion-{hostname}">
                        <div class="accordion-body">
                            <p><strong>Timestamp:</strong> {cmd_result['timestamp']}</p>
"""

            if cmd_result["status"] == "success":
                html_content += f"""
                            <p><strong>Output:</strong></p>
                            <div class="command-output">{cmd_result['output']}</div>
"""
            else:
                html_content += f"""
                            <div class="alert alert-danger">
                                <strong>Error:</strong> {cmd_result['output']}
                            </div>
"""

            html_content += """
                        </div>
                    </div>
                </div>
"""

        html_content += """
            </div>
        </div>
"""

    html_content += """
        <!-- Footer -->
        <div class="text-center mt-5 mb-3 text-muted">
            <p>Generated by Netmiko Device Command Collector v3.0
            </p>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js">
    </script>
</body>
</html>
"""

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html_content)

    logger.info("HTML report saved to %s", output_file)


def save_to_excel(results: List[Dict[str, str]], output_file: str) -> None:
    """
    Save command outputs to an Excel file with beautiful formatting.

    Args:
        results: List of result dictionaries
        output_file: Path to the output Excel file
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed. Cannot create Excel file.")
        logger.info("Install with: pip install openpyxl")
        return

    if not results:
        logger.warning("No results to save")
        return

    # Create workbook
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Calculate statistics
    devices = set(r["hostname"] for r in results)
    total_commands = len(results)
    successful = sum(1 for r in results if r["status"] == "success")
    failed = total_commands - successful

    # Write summary
    ws_summary["A1"] = "Network Device Command Collection Report"
    ws_summary["A1"].font = Font(bold=True, size=16, color="4F46E5")
    ws_summary.merge_cells("A1:B1")

    ws_summary["A3"] = "Generated:"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws_summary["A5"] = "Metric"
    ws_summary["B5"] = "Value"
    ws_summary["A5"].fill = header_fill
    ws_summary["B5"].fill = header_fill
    ws_summary["A5"].font = header_font
    ws_summary["B5"].font = header_font

    summary_data = [
        ("Total Devices", len(devices)),
        ("Total Commands", total_commands),
        ("Successful", successful),
        ("Failed", failed),
        ("Success Rate", f"{(successful/total_commands*100):.1f}%"),
    ]

    for idx, (metric, value) in enumerate(summary_data, start=6):
        ws_summary[f"A{idx}"] = metric
        ws_summary[f"B{idx}"] = value
        if metric == "Successful":
            ws_summary[f"B{idx}"].font = Font(color="10B981", bold=True)
        elif metric == "Failed" and failed > 0:
            ws_summary[f"B{idx}"].font = Font(color="EF4444", bold=True)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20

    # Detailed results sheet
    ws_details = wb.create_sheet("Detailed Results")

    headers = ["Timestamp", "Hostname", "IP Address", "Command", "Output", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, result in enumerate(results, start=2):
        ws_details.cell(row=row_num, column=1, value=result["timestamp"])
        ws_details.cell(row=row_num, column=2, value=result["hostname"])
        ws_details.cell(row=row_num, column=3, value=result["ip_address"])
        ws_details.cell(row=row_num, column=4, value=result["command"])
        ws_details.cell(row=row_num, column=5, value=result["output"])

        status_cell = ws_details.cell(row=row_num, column=6, value=result["status"])

        if result["status"] == "success":
            status_cell.fill = PatternFill(
                start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
            )
            status_cell.font = Font(color="059669", bold=True)
        else:
            status_cell.fill = PatternFill(
                start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
            )
            status_cell.font = Font(color="DC2626", bold=True)

    # Adjust column widths
    ws_details.column_dimensions["A"].width = 20
    ws_details.column_dimensions["B"].width = 20
    ws_details.column_dimensions["C"].width = 15
    ws_details.column_dimensions["D"].width = 30
    ws_details.column_dimensions["E"].width = 50
    ws_details.column_dimensions["F"].width = 12

    # Freeze header row
    ws_details.freeze_panes = "A2"

    # Save workbook
    wb.save(output_file)
    logger.info("Excel report saved to %s", output_file)


def save_results(
    results: List[Dict[str, str]], output_file: str, formats: Optional[List[str]] = None
) -> None:
    """
    Save results in multiple formats.

    Args:
        results: List of result dictionaries
        output_file: Base output filename (extension will be replaced)
        formats: List of formats to save ('csv', 'json', 'html', 'markdown', 'excel')
    """
    if formats is None:
        formats = ["csv"]

    base_name = output_file.rsplit(".", 1)[0]

    for fmt in formats:
        if fmt == "csv":
            save_to_csv(results, f"{base_name}.csv")
        elif fmt == "json":
            save_to_json(results, f"{base_name}.json")
        elif fmt == "html":
            save_to_html(results, f"{base_name}.html")
        elif fmt in ("markdown", "md"):
            save_to_markdown(results, f"{base_name}.md")
        elif fmt in ("excel", "xlsx"):
            save_to_excel(results, f"{base_name}.xlsx")
        else:
            logger.warning("Unknown format: %s", fmt)


# ====================================================================================
# TYPER CLI APPLICATION
# ====================================================================================

# Initialize Typer app
app = typer.Typer(
    name="netmiko-collector",
    help="🚀 Network Device Command Collector using Netmiko",
    add_completion=True,
    rich_markup_mode="rich",
    no_args_is_help=True,
)


# ====================================================================================
# CLI COMMANDS
# ====================================================================================


@app.command(name="run", help="🚀 Run command collection on network devices")
def run_collection(
    devices: Annotated[
        Optional[str], typer.Option("--devices", "-d", help="Path to devices CSV file")
    ] = None,
    commands: Annotated[
        Optional[str], typer.Option("--commands", "-c", help="Path to commands text file")
    ] = None,
    output: Annotated[
        str, typer.Option("--output", "-o", help="Output file base name")
    ] = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
    output_format: Annotated[
        Optional[List[str]], typer.Option("--format", "-f", help="Output format(s)")
    ] = None,
    username: Annotated[
        Optional[str], typer.Option("--username", "-u", help="SSH username")
    ] = None,
    ssh_config: Annotated[
        Optional[str], typer.Option("--ssh-config", "-s", help="SSH config file path")
    ] = None,
    device_type: Annotated[
        Optional[str], typer.Option("--device-type", help="Default device type")
    ] = None,
    workers: Annotated[
        Optional[int], typer.Option("--workers", "-w", help="Number of concurrent workers")
    ] = None,
    connection_timeout: Annotated[
        Optional[int], typer.Option("--connection-timeout", help="Connection timeout in seconds")
    ] = None,
    command_timeout: Annotated[
        Optional[int], typer.Option("--command-timeout", help="Command timeout in seconds")
    ] = None,
    enable_mode: Annotated[bool, typer.Option("--enable-mode", help="Enter enable mode")] = False,
    enable_password: Annotated[
        Optional[str], typer.Option("--enable-password", help="Enable mode password")
    ] = None,
    enable_session_logging: Annotated[
        bool, typer.Option("--enable-session-logging", help="Enable session logging")
    ] = False,
    no_strip_whitespace: Annotated[
        bool, typer.Option("--no-strip-whitespace", help="Don't strip whitespace")
    ] = False,
    no_retry: Annotated[bool, typer.Option("--no-retry", help="Disable retry on failure")] = False,
):
    """
    Run command collection on network devices.

    Examples:
        netmiko-collector run -d devices.csv -c commands.txt
        netmiko-collector run -d devices.csv -c commands.txt -f json html
        netmiko-collector run -d devices.csv -c commands.txt -w 10
    """
    # Load configuration
    config = load_config()

    # Handle default output_format value
    if output_format is None:
        output_format = ["csv"]

    # Interactive mode if no files provided
    if not devices or not commands:
        typer.echo("📋 Interactive Mode\n")

        # Get devices file
        if not devices:
            devices = get_or_create_file("devices.csv", "devices CSV")
            if not devices:
                typer.secho("❌ Devices file is required", fg=typer.colors.RED)
                raise typer.Exit(1)

        # Get commands file
        if not commands:
            commands = get_or_create_file("commands.txt", "commands")
            if not commands:
                typer.secho("❌ Commands file is required", fg=typer.colors.RED)
                raise typer.Exit(1)

    # Get credentials
    if not username:
        username = typer.prompt("Enter SSH username")

    password = getpass("Enter SSH password: ")

    # Ask about enable mode if config has it enabled
    if config.get("enable_mode", False) and not enable_mode:
        if typer.confirm("Enter enable mode?", default=True):
            enable_mode = True
            enable_password = getpass(
                "Enter enable password (or press Enter to use SSH password): "
            )
            if not enable_password:
                enable_password = password

    # Apply configuration defaults
    if device_type is None:
        device_type = config["default_device_type"]
    if workers is None:
        workers = config["max_workers"]
    if connection_timeout is None:
        connection_timeout = config["connection_timeout"]
    if command_timeout is None:
        command_timeout = config["command_timeout"]

    strip_whitespace = not no_strip_whitespace
    retry_on_failure = not no_retry

    # Process formats
    output_formats = []
    for fmt in output_format:
        if fmt in ["markdown", "md"]:
            output_formats.append("markdown")
        elif fmt in ["excel", "xlsx"]:
            output_formats.append("excel")
        elif fmt == "all":
            output_formats = ["csv", "json", "html", "markdown", "excel"]
            break
        else:
            output_formats.append(fmt)

    # Load devices and commands
    try:
        devices_list = load_devices(devices, device_type)
        commands_list = load_commands(commands)

        typer.secho(
            f"\n✅ Loaded {len(devices_list)} device(s) and {len(commands_list)} command(s)",
            fg=typer.colors.GREEN,
        )

        # Process devices
        all_results = process_devices_parallel(
            devices_list,
            commands_list,
            username,
            password,
            ssh_config,
            strip_whitespace,
            connection_timeout,
            command_timeout,
            workers,
            enable_session_logging,
            enable_mode,
            enable_password,
            retry_on_failure,
            show_progress=True,
        )

        # Save results
        save_results(all_results, output, output_formats)

        typer.secho("\n✅ Collection completed successfully!", fg=typer.colors.GREEN, bold=True)

    except Exception as e:
        typer.secho(f"❌ Error: {e}", fg=typer.colors.RED)
        raise typer.Exit(1)


@app.command(name="edit", help="✏️ Edit devices or commands files")
def edit_files(
    file_type: Annotated[
        str, typer.Argument(help="File type to edit: 'devices' or 'commands'")
    ] = "devices",
):
    """
    Edit devices.csv or commands.txt files.

    Examples:
        netmiko-collector edit devices
        netmiko-collector edit commands
    """
    if file_type.lower() in ["devices", "device", "dev", "d"]:
        # Edit devices.csv
        devices_file = get_or_create_file("devices.csv", "devices CSV")

        if devices_file:
            typer.secho(f"📝 Opening {devices_file}...", fg=typer.colors.CYAN)
            if open_file_in_editor(devices_file):
                typer.secho("✅ Editor closed", fg=typer.colors.GREEN)
        else:
            typer.secho("ℹ️  Operation cancelled", fg=typer.colors.YELLOW)

    elif file_type.lower() in ["commands", "command", "cmd", "c"]:
        # Edit commands.txt with autocomplete
        commands_file = get_or_create_file("commands.txt", "commands")

        if commands_file:
            if PROMPT_TOOLKIT_AVAILABLE:
                config = load_config()
                device_type = config.get("default_device_type", "cisco_ios")

                typer.secho(
                    f"📝 Opening {commands_file} with autocomplete...", fg=typer.colors.CYAN
                )
                typer.secho(f"🔧 Device type: {device_type}", fg=typer.colors.CYAN)
                typer.secho(
                    "💡 Tip: Press Tab for suggestions, type 'done' when finished",
                    fg=typer.colors.YELLOW,
                )

                if edit_commands_with_autocomplete(commands_file, device_type):
                    typer.secho("✅ Commands file updated", fg=typer.colors.GREEN)
                else:
                    typer.secho("⚠️  No changes made", fg=typer.colors.YELLOW)
            else:
                typer.secho(
                    "⚠️  prompt_toolkit not available - using basic editor", fg=typer.colors.YELLOW
                )
                if open_file_in_editor(commands_file):
                    typer.secho("✅ Editor closed", fg=typer.colors.GREEN)
        else:
            typer.secho("ℹ️  Operation cancelled", fg=typer.colors.YELLOW)
    else:
        typer.secho(f"❌ Unknown file type: {file_type}", fg=typer.colors.RED)
        typer.secho("Use 'devices' or 'commands'", fg=typer.colors.YELLOW)
        raise typer.Exit(1)


@app.command(name="view", help="👁️  View output files")
def view_outputs():
    """
    View generated output files interactively.

    Example:
        netmiko-collector view
    """
    output_files = list_output_files()

    if not output_files:
        typer.secho("⚠️  No output files found", fg=typer.colors.YELLOW)
        return

    typer.secho("\n📂 Available Output Files:", fg=typer.colors.CYAN, bold=True)
    for idx, file in enumerate(output_files, 1):
        file_obj = Path(file)
        file_type = file_obj.suffix.upper().replace(".", "")
        typer.secho(f"  {idx}. {file_obj.name} ({file_type})", fg=typer.colors.GREEN)

    try:
        choice = typer.prompt(
            "\nEnter file number to view (or press Enter to cancel)", default="", show_default=False
        )
        if choice:
            file_idx = int(choice) - 1
            if 0 <= file_idx < len(output_files):
                view_output_file(output_files[file_idx])
            else:
                typer.secho("❌ Invalid file number", fg=typer.colors.RED)
    except (ValueError, KeyboardInterrupt):
        typer.secho("\nℹ️  Cancelled", fg=typer.colors.YELLOW)


# Config subcommand group
config_app = typer.Typer(help="⚙️  Manage configuration settings")
app.add_typer(config_app, name="config")


@config_app.command(name="show", help="Display current configuration")
def config_show():
    """Show current configuration settings."""
    config = load_config()

    typer.secho("\n⚙️  Current Configuration", fg=typer.colors.CYAN, bold=True)
    typer.secho("=" * 50, fg=typer.colors.CYAN)

    for key, value in config.items():
        typer.secho(f"  {key}: ", nl=False, fg=typer.colors.YELLOW)
        typer.secho(f"{value}", fg=typer.colors.GREEN)


@config_app.command(name="set", help="Set configuration values interactively")
def config_set():
    """Set configuration values interactively."""
    config = load_config()

    typer.secho("\n⚙️  Configuration Editor", fg=typer.colors.CYAN, bold=True)
    typer.secho("Press Enter to keep current value\n", fg=typer.colors.YELLOW)

    # Device type
    new_device_type = typer.prompt(
        f"Default device type [{config['default_device_type']}]",
        default=config["default_device_type"],
    )
    config["default_device_type"] = new_device_type

    # Strip whitespace
    config["strip_whitespace"] = typer.confirm(
        "Strip whitespace?", default=config["strip_whitespace"]
    )

    # Max workers
    config["max_workers"] = typer.prompt("Max workers", type=int, default=config["max_workers"])

    # Connection timeout
    config["connection_timeout"] = typer.prompt(
        "Connection timeout (seconds)", type=int, default=config["connection_timeout"]
    )

    # Command timeout
    config["command_timeout"] = typer.prompt(
        "Command timeout (seconds)", type=int, default=config["command_timeout"]
    )

    # Enable session logging
    config["enable_session_logging"] = typer.confirm(
        "Enable session logging?", default=config.get("enable_session_logging", False)
    )

    # Enable mode
    config["enable_mode"] = typer.confirm(
        "Enter enable mode on connect?", default=config.get("enable_mode", False)
    )

    # Retry on failure
    config["retry_on_failure"] = typer.confirm(
        "Retry on connection failure?", default=config.get("retry_on_failure", True)
    )

    # Save configuration
    save_config(config)
    typer.secho("\n✅ Configuration saved successfully!", fg=typer.colors.GREEN, bold=True)


@config_app.command(name="reset", help="Reset configuration to defaults")
def config_reset():
    """Reset configuration to default values."""
    if typer.confirm("⚠️  Reset all settings to defaults?", default=False):
        config = DEFAULT_CONFIG.copy()
        save_config(config)
        typer.secho("✅ Configuration reset to defaults", fg=typer.colors.GREEN)
    else:
        typer.secho("ℹ️  Cancelled", fg=typer.colors.YELLOW)


# Sample file creation subcommand group
sample_app = typer.Typer(help="📝 Create sample files")
app.add_typer(sample_app, name="sample")


@sample_app.command(name="devices", help="Create sample devices.csv file")
def sample_devices():
    """Create a sample devices.csv file."""
    sample_content = """hostname,ip_address,device_type
router1,192.168.1.1,cisco_ios
switch1,192.168.1.2,cisco_ios
router2,10.0.0.1,cisco_xe
"""

    file_path = Path("devices.csv")

    if file_path.exists():
        if not typer.confirm(f"⚠️  {file_path} already exists. Overwrite?", default=False):
            typer.secho("ℹ️  Cancelled", fg=typer.colors.YELLOW)
            return

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(sample_content)
        typer.secho(f"✅ Created sample {file_path}", fg=typer.colors.GREEN)
    except Exception as e:
        typer.secho(f"❌ Failed to create file: {e}", fg=typer.colors.RED)


@sample_app.command(name="commands", help="Create sample commands.txt file")
def sample_commands():
    """Create a sample commands.txt file."""
    sample_content = """# Sample network commands
show version
show ip interface brief
show running-config | include hostname
show inventory
"""

    file_path = Path("commands.txt")

    if file_path.exists():
        if not typer.confirm(f"⚠️  {file_path} already exists. Overwrite?", default=False):
            typer.secho("ℹ️  Cancelled", fg=typer.colors.YELLOW)
            return

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(sample_content)
        typer.secho(f"✅ Created sample {file_path}", fg=typer.colors.GREEN)
    except Exception as e:
        typer.secho(f"❌ Failed to create file: {e}", fg=typer.colors.RED)


if __name__ == "__main__":
    app()
