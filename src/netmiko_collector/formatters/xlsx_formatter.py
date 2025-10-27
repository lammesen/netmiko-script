"""XLSX formatter for command execution results."""

from typing import List

from ..models import ExecutionResult
from .base import BaseFormatter


class XLSXFormatter(BaseFormatter):
    """Format execution results as Excel XLSX file."""

    def format(self, results: List[ExecutionResult]) -> str:
        """Format execution results as XLSX (returns binary data as string).
        
        Note: This method returns a string representation. Use write_to_file()
        for actual XLSX file creation.
        
        Args:
            results: List of ExecutionResult objects
            
        Returns:
            String indicating XLSX format (actual binary written by write_to_file)
        """
        return f"XLSX format with {len(results)} results"
    
    def write_to_file(self, results: List[ExecutionResult], filepath) -> None:
        """Write formatted results to an XLSX file.
        
        Args:
            results: List of ExecutionResult objects to format
            filepath: Path to output file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX output. "
                "Install with: pip install openpyxl"
            )
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Command Results"
        
        # Define headers
        headers = ["Hostname", "Command", "Status", "Output", "Error", "Duration (ms)"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for result in results:
            ws.append([
                result.hostname,
                result.command.text,
                result.status.value,
                result.output or "",
                result.error or "",
                int(result.duration * 1000) if result.duration is not None else ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass  # Skip cells that can't be measured
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(str(filepath))
