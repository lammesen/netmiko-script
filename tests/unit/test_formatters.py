"""Tests for output formatters."""

import json
from datetime import datetime
from pathlib import Path

import pytest

from src.netmiko_collector.models import Command, Device, DeviceType, ExecutionResult, ExecutionStatus
from src.netmiko_collector.formatters import (
    CSVFormatter,
    JSONFormatter,
    YAMLFormatter,
    HTMLFormatter,
    XLSXFormatter,
)


@pytest.fixture
def sample_results():
    """Create sample execution results for testing."""
    return [
        ExecutionResult(
            device=Device(hostname="router1", device_type=DeviceType.CISCO_IOS, username="admin"),
            command=Command(command="show version"),
            status=ExecutionStatus.SUCCESS,
            output="Cisco IOS Software, Version 15.2",
            error="",
            duration=1.5,
            timestamp=datetime(2025, 1, 1, 12, 0, 0)
        ),
        ExecutionResult(
            device=Device(hostname="router2", device_type=DeviceType.CISCO_IOS, username="admin"),
            command=Command(command="show interfaces"),
            status=ExecutionStatus.FAILED,
            output="",
            error="Authentication failed",
            duration=0.5,
            timestamp=datetime(2025, 1, 1, 12, 0, 1)
        ),
        ExecutionResult(
            device=Device(hostname="router3", device_type=DeviceType.CISCO_IOS, username="admin"),
            command=Command(command="show ip route"),
            status=ExecutionStatus.SUCCESS,
            output="Gateway of last resort is 10.0.0.1",
            error="",
            duration=0.0,
            timestamp=datetime.now()
        ),
    ]


class TestCSVFormatter:
    """Tests for CSV formatter."""

    def test_format_basic(self, sample_results):
        """Test basic CSV formatting."""
        formatter = CSVFormatter()
        output = formatter.format(sample_results)
        
        lines = output.strip().split("\n")
        assert len(lines) == 4  # Header + 3 results
        # Check header (strip any trailing whitespace)
        assert lines[0].strip() == "hostname,command,status,output,error,duration_ms"
        assert "router1" in lines[1]
        assert "show version" in lines[1]
        assert "success" in lines[1]
        assert "1500" in lines[1]  # 1.5s = 1500ms
    
    def test_format_empty(self):
        """Test formatting empty results."""
        formatter = CSVFormatter()
        output = formatter.format([])
        
        lines = output.strip().split("\n")
        assert len(lines) == 1  # Just header
        assert lines[0].strip() == "hostname,command,status,output,error,duration_ms"
    
    def test_format_with_error(self, sample_results):
        """Test formatting results with errors."""
        formatter = CSVFormatter()
        output = formatter.format(sample_results)
        
        assert "Authentication failed" in output
        assert "failed" in output
    
    def test_format_with_none_duration(self, sample_results):
        """Test formatting with zero duration."""
        formatter = CSVFormatter()
        output = formatter.format(sample_results)
        
        lines = output.strip().split("\n")
        # Last result has 0.0 duration
        assert "0" in lines[3]  # 0 ms
    
    def test_write_to_file(self, sample_results, tmp_path):
        """Test writing CSV to file."""
        formatter = CSVFormatter()
        filepath = tmp_path / "output.csv"
        
        formatter.write_to_file(sample_results, filepath)
        
        assert filepath.exists()
        content = filepath.read_text()
        assert "hostname,command,status" in content
        assert "router1" in content


class TestJSONFormatter:
    """Tests for JSON formatter."""

    def test_format_basic(self, sample_results):
        """Test basic JSON formatting."""
        formatter = JSONFormatter()
        output = formatter.format(sample_results)
        
        data = json.loads(output)
        assert len(data) == 3
        assert data[0]["hostname"] == "router1"
        assert data[0]["command"] == "show version"
        assert data[0]["status"] == "success"
        assert data[0]["duration_ms"] == 1500
    
    def test_format_with_indent(self, sample_results):
        """Test JSON formatting with custom indent."""
        formatter = JSONFormatter(indent=4)
        output = formatter.format(sample_results)
        
        assert "    " in output  # 4-space indent
        data = json.loads(output)
        assert len(data) == 3
    
    def test_format_empty(self):
        """Test formatting empty results."""
        formatter = JSONFormatter()
        output = formatter.format([])
        
        data = json.loads(output)
        assert data == []
    
    def test_format_with_none_values(self, sample_results):
        """Test formatting with empty values."""
        formatter = JSONFormatter()
        output = formatter.format(sample_results)
        
        data = json.loads(output)
        # Second result has empty output (empty string, not None)
        assert data[1]["output"] == ""
        # Third result has 0.0 duration
        assert data[2]["duration_ms"] == 0
    
    def test_format_timestamp(self, sample_results):
        """Test timestamp formatting."""
        formatter = JSONFormatter()
        output = formatter.format(sample_results)
        
        data = json.loads(output)
        assert data[0]["timestamp"] == "2025-01-01T12:00:00"


class TestYAMLFormatter:
    """Tests for YAML formatter."""

    def test_format_basic(self, sample_results):
        """Test basic YAML formatting."""
        pytest.importorskip("yaml")
        import yaml
        
        formatter = YAMLFormatter()
        output = formatter.format(sample_results)
        
        data = yaml.safe_load(output)
        assert len(data) == 3
        assert data[0]["hostname"] == "router1"
        assert data[0]["command"] == "show version"
        assert data[0]["status"] == "success"
    
    def test_format_missing_yaml(self, sample_results, monkeypatch):
        """Test error when PyYAML is not installed."""
        # Mock yaml import to fail
        import sys
        monkeypatch.setitem(sys.modules, "yaml", None)
        
        formatter = YAMLFormatter()
        with pytest.raises(ImportError, match="PyYAML is required"):
            formatter.format(sample_results)
    
    def test_format_empty(self):
        """Test formatting empty results."""
        pytest.importorskip("yaml")
        import yaml
        
        formatter = YAMLFormatter()
        output = formatter.format([])
        
        data = yaml.safe_load(output)
        assert data == []


class TestHTMLFormatter:
    """Tests for HTML formatter."""

    def test_format_basic(self, sample_results):
        """Test basic HTML formatting."""
        formatter = HTMLFormatter()
        output = formatter.format(sample_results)
        
        assert "<!DOCTYPE html>" in output
        assert "<html>" in output
        assert "<table>" in output
        assert "router1" in output
        assert "show version" in output
    
    def test_format_with_status_classes(self, sample_results):
        """Test HTML status classes."""
        formatter = HTMLFormatter()
        output = formatter.format(sample_results)
        
        assert "class='success'" in output
        assert "class='failed'" in output
    
    def test_format_escape_html(self, sample_results):
        """Test HTML escaping."""
        # Add result with HTML characters
        sample_results.append(
            ExecutionResult(
                device=Device(hostname="router4", device_type=DeviceType.CISCO_IOS, username="admin"),
                command=Command(command='<script>alert("test")</script>'),
                status=ExecutionStatus.SUCCESS,
                output='<html>&test',
                error="",
                duration=1.0,
                timestamp=datetime.now()
            )
        )
        
        formatter = HTMLFormatter()
        output = formatter.format(sample_results)
        
        assert "&lt;script&gt;" in output
        assert "&amp;test" in output
        assert "<script>" not in output  # Should be escaped
    
    def test_format_empty(self):
        """Test formatting empty results."""
        formatter = HTMLFormatter()
        output = formatter.format([])
        
        assert "<table>" in output
        assert "Total Results: 0" in output
    
    def test_format_duration_display(self, sample_results):
        """Test duration display in HTML."""
        formatter = HTMLFormatter()
        output = formatter.format(sample_results)
        
        assert "1500" in output  # 1.5s = 1500ms
        assert "500" in output   # 0.5s = 500ms
        assert "0" in output     # 0.0 duration


class TestXLSXFormatter:
    """Tests for XLSX formatter."""

    def test_format_returns_description(self, sample_results):
        """Test format method returns description."""
        formatter = XLSXFormatter()
        output = formatter.format(sample_results)
        
        assert "XLSX format" in output
        assert "3 results" in output
    
    def test_write_to_file_basic(self, sample_results, tmp_path):
        """Test writing XLSX to file."""
        pytest.importorskip("openpyxl")
        import openpyxl
        
        formatter = XLSXFormatter()
        filepath = tmp_path / "output.xlsx"
        
        formatter.write_to_file(sample_results, filepath)
        
        assert filepath.exists()
        
        # Verify content
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Check headers
        assert ws['A1'].value == "Hostname"
        assert ws['B1'].value == "Command"
        assert ws['C1'].value == "Status"
        
        # Check data
        assert ws['A2'].value == "router1"
        assert ws['B2'].value == "show version"
        assert ws['C2'].value == "success"
    
    def test_write_to_file_missing_openpyxl(self, sample_results, tmp_path, monkeypatch):
        """Test error when openpyxl is not installed."""
        # Mock openpyxl import to fail
        import sys
        monkeypatch.setitem(sys.modules, "openpyxl", None)
        
        formatter = XLSXFormatter()
        filepath = tmp_path / "output.xlsx"
        
        with pytest.raises(ImportError, match="openpyxl is required"):
            formatter.write_to_file(sample_results, filepath)
    
    def test_write_to_file_with_errors(self, sample_results, tmp_path):
        """Test writing XLSX with error results."""
        pytest.importorskip("openpyxl")
        import openpyxl
        
        formatter = XLSXFormatter()
        filepath = tmp_path / "output.xlsx"
        
        formatter.write_to_file(sample_results, filepath)
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Check error in second result
        assert ws['E3'].value == "Authentication failed"
        assert ws['C3'].value == "failed"
